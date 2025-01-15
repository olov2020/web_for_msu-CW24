from __future__ import annotations  # Поддержка строковых аннотаций

from datetime import datetime
from io import BytesIO
from typing import TYPE_CHECKING

import pandas as pd
import pytz
from flask import send_file
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Border, Side
from sqlalchemy.sql import text

if TYPE_CHECKING:
    # Импортируем сервисы только для целей аннотации типов
    pass


class BackupService:
    def __init__(self, db):
        self.db = db

    def export_db(self):
        # Создаём объект Excel в памяти с корректным управлением контекстом записи
        output = BytesIO()

        # Используем контекстный менеджер для ExcelWriter
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Получаем список всех таблиц в базе данных
            inspector = self.db.inspect(self.db.engine)
            table_names = inspector.get_table_names()
            dfs = {}  # Словарь для хранения DataFrame'ов

            # Экспортируем каждую таблицу в отдельный лист Excel
            for table_name in table_names:
                # Читаем данные из таблицы в DataFrame
                query = self.db.session.execute(text(f'SELECT * FROM \"{table_name}\"'))
                df = pd.DataFrame(query.fetchall(), columns=query.keys())
                dfs[table_name] = df  # Сохраняем DataFrame в словаре
                # Проверяем, не пустой ли DataFrame перед записью
                if not df.empty:
                    df.to_excel(writer, sheet_name=table_name, index=False)

        # Перемещаем указатель потока в начало
        output.seek(0)

        wb = load_workbook(output)

        if "formula" in wb.sheetnames:
            ws_formula = wb["formula"]
            if "formula" in dfs:  # Проверяем, есть ли DataFrame для "formula"
                df_formula = dfs["formula"]  # Получаем DataFrame из словаря
                self.apply_alternating_colors(ws_formula, df_formula)
        if "schedule" in wb.sheetnames:
            ws_formula = wb["schedule"]
            if "schedule" in dfs:  # Проверяем, есть ли DataFrame для "formula"
                df_formula = dfs["schedule"]  # Получаем DataFrame из словаря
                self.apply_alternating_colors(ws_formula, df_formula)

        temp_output = BytesIO()
        wb.save(temp_output)
        temp_output.seek(0)

        date = datetime.now(tz=pytz.timezone('Europe/Moscow')).date().strftime('%d.%m.%Y')
        # Отправляем файл пользователю
        return send_file(
            temp_output,
            download_name=f'emsh_data_{date}.xlsx',
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    def apply_alternating_colors(self, ws, df):
        """Применяет чередование цветов к строкам на листе Excel."""
        if df.empty:
            return  # Ничего не делаем, если DataFrame пустой

        fill1 = PatternFill(fill_type='solid', start_color='E0F2F7')  # Светло-голубой
        fill2 = PatternFill(fill_type='solid', start_color='FFFFFF')  # Белый

        course_id_column = df['course_id'].tolist()

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        current_fill = fill1
        for row_index, course_id in enumerate(course_id_column, start=2):
            if row_index > 2 and course_id != course_id_column[row_index - 3]:
                current_fill = fill2 if current_fill == fill1 else fill1
            for cell in ws[f'A{row_index}:AA{row_index}'][0]:  # Применяем заливку ко всей строке (от A до D)
                cell.fill = current_fill
                cell.border = thin_border
